from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl

from juval.domain.costs import CostInputs, FeeInputs
from juval.domain.product import Identification, Product
from juval.domain.provenance import FieldValue, VerificationStatus
from juval.domain.risk import RiskFlag, RiskProfile, RiskStatus, RiskType, Severity
from juval.domain.sourcing_record import SourcingRecord
from juval.infrastructure.excel.exporter import HEADERS, export_excel

NOW = datetime(2026, 1, 1, tzinfo=timezone.utc)


def _record():
    asin = FieldValue.verified("B0EXPORTED1", source="s", method="m", retrieved_at=NOW)
    identification = Identification(asin=asin, marketplace="US", supplier_sku="SUP-EXPORT")
    product = Product(identification=identification)
    costs = CostInputs(cog=Decimal("3"))
    flag = RiskFlag(
        risk_type=RiskType.HAZMAT, status=RiskStatus.ABSENT, verification_status=VerificationStatus.VERIFIED,
        severity=Severity.NONE, source="s", timestamp=NOW,
    )
    return SourcingRecord(record_ref="row_1:SUP-EXPORT", product=product, costs=costs, risk=RiskProfile(flags=(flag,)))


def test_export_writes_header_row(tmp_path):
    path = tmp_path / "out.xlsx"
    export_excel([_record()], path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == HEADERS


def test_export_never_collapses_a_field_value_to_a_single_cell(tmp_path):
    path = tmp_path / "out.xlsx"
    export_excel([_record()], path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row, data_row = list(ws.iter_rows(values_only=True))[:2]
    row = dict(zip(header_row, data_row))
    assert row["asin"] == "B0EXPORTED1"
    assert row["asin_status"] == "VERIFIED"  # value and status are separate columns


def test_export_includes_identification_costs_risk(tmp_path):
    path = tmp_path / "out.xlsx"
    export_excel([_record()], path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row, data_row = list(ws.iter_rows(values_only=True))[:2]
    row = dict(zip(header_row, data_row))
    assert row["record_ref"] == "row_1:SUP-EXPORT"
    assert row["supplier_sku"] == "SUP-EXPORT"
    assert row["cog"] == Decimal("3")
    assert row["hazmat_status"] == "ABSENT"


def test_export_includes_max_cog_targets(tmp_path):
    path = tmp_path / "out.xlsx"

    from juval.processing.profitability import compute_profitability

    record = _record()
    costs = record.costs
    fees = FeeInputs(referral_fee=Decimal("3"), referral_fee_rate=Decimal("0.15"), fulfillment_fee=Decimal("2"))
    selling_price = FieldValue.verified(Decimal("19.99"), source="s", method="m", retrieved_at=NOW)
    profitability = compute_profitability(
        selling_price=selling_price, costs=costs, fees=fees,
        target_profit=Decimal("5"), target_roi=Decimal("0.3"), now=NOW,
    )
    record = record.with_profitability(profitability)

    export_excel([record], path)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row, data_row = list(ws.iter_rows(values_only=True))[:2]
    row = dict(zip(header_row, data_row))
    # openpyxl round-trips Decimal through float, so compare with a small
    # tolerance rather than exact equality (see _EPSILON in data_quality.py).
    assert abs(Decimal(str(row["max_cog_target_profit"])) - profitability.max_cog_target_profit.value) < Decimal("0.01")
    assert row["max_cog_target_profit_status"] == "VERIFIED"
    assert abs(Decimal(str(row["max_cog_target_roi"])) - profitability.max_cog_target_roi.value) < Decimal("0.01")
    assert row["max_cog_target_roi_status"] == "VERIFIED"


def test_export_empty_records_still_writes_headers(tmp_path):
    path = tmp_path / "empty.xlsx"
    export_excel([], path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [HEADERS]
