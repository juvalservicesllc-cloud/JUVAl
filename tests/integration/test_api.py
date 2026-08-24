"""Integration tests for interfaces/api/main.py (Fase 4A).

Uses FastAPI's TestClient (httpx under the hood) -- no real server, no
network socket, same "invoke directly" philosophy as test_cli.py. The
main path drives the real Excel fixture through the real Core
(run_pipeline -> processing -> domain), never a mock, per the Fase 4A
brief ("no crear un mock que sustituya completamente al Core").
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from juval.infrastructure.logging.sqlite_execution_run_store import SqliteExecutionRunStore
from juval.interfaces.api.main import app

FIXTURE = Path(__file__).parent.parent / "fixtures" / "sample_sourcing_TEST_DATA.xlsx"

_THRESHOLDS = json.dumps(
    {
        "target_profit": "5",
        "target_roi": "0.3",
        "minimum_estimated_monthly_sales": 0,
        "maximum_risk_severity": "LOW",
    }
)
_FEES = json.dumps({"referral_fee": "3", "referral_fee_rate": "0.15", "fulfillment_fee": "2"})


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("JUVAL_RUN_STORAGE_DIR", str(tmp_path))
    monkeypatch.setenv("JUVAL_EXECUTION_DB_PATH", str(tmp_path / "execution_runs.db"))
    monkeypatch.delenv("JUVAL_MAX_UPLOAD_BYTES", raising=False)
    return TestClient(app)


def _upload(client, *, thresholds=_THRESHOLDS, fees=_FEES, persist=None, filename="sample.xlsx", content=None):
    data = {"thresholds": thresholds, "fees": fees}
    if persist is not None:
        data["persist"] = persist
    file_bytes = content if content is not None else FIXTURE.read_bytes()
    return client.post(
        "/api/v1/runs",
        files={"file": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data=data,
    )


def _batch(client, files, *, persist=None):
    data = {"thresholds": _THRESHOLDS, "fees": _FEES}
    if persist is not None:
        data["persist"] = persist
    return client.post(
        "/api/v1/batches",
        files=[("files", (name, content, content_type)) for name, content, content_type in files],
        data=data,
    )


def _fixture_as_csv() -> bytes:
    """Re-emit the XLSX fixture as CSV so both formats carry identical data."""
    import csv
    import io

    import openpyxl

    workbook = openpyxl.load_workbook(FIXTURE, data_only=True, read_only=True)
    try:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer, lineterminator="\r\n")
        for row in workbook.worksheets[0].iter_rows(values_only=True):
            writer.writerow(["" if cell is None else cell for cell in row])
    finally:
        workbook.close()
    return buffer.getvalue().encode("utf-8")


def test_endpoint_exists(client):
    resp = _upload(client)
    assert resp.status_code != 404


def test_batch_processes_two_files_and_preserves_ordered_child_runs(client):
    payload = FIXTURE.read_bytes()
    response = _batch(client, [("first.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("second.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")], persist="true")
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "PARTIAL_SUCCESS"
    assert [item["filename"] for item in body["files"]] == ["first.xlsx", "second.xlsx"]
    assert all(item["execution_id"] for item in body["files"])
    loaded = client.get(f"/api/v1/batches/{body['batch_id']}")
    assert loaded.status_code == 200
    assert loaded.json()["files"] == body["files"]


def test_batch_rejects_overflow_without_processing(client):
    payload = FIXTURE.read_bytes()
    files = [(f"file-{index}.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") for index in range(11)]
    response = _batch(client, files)
    assert response.status_code == 422
    assert "at most 10" in response.json()["detail"]


def test_batch_isolates_unimportable_file_and_keeps_valid_sibling(client):
    """A CSV without the required columns fails on its own; its sibling still runs."""
    payload = FIXTURE.read_bytes()
    response = _batch(client, [("valid.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("bad.csv", b"a,b\n1,2", "text/csv")])
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "PARTIAL_SUCCESS"
    assert [item["status"] for item in body["files"]] == ["PARTIAL_SUCCESS", "FAILED"]
    assert body["files"][0]["execution_id"]
    assert body["files"][1]["records_processed"] == 0


def test_batch_rejects_unsupported_file_type_by_name(client):
    payload = FIXTURE.read_bytes()
    response = _batch(client, [("valid.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), ("notes.pdf", b"%PDF-1.4", "application/pdf")])
    assert response.status_code == 200
    body = response.json()
    files = body["files"]
    assert [item["status"] for item in files] == ["PARTIAL_SUCCESS", "REJECTED"]
    assert files[1]["filename"] == "notes.pdf"
    assert files[1]["execution_id"] is None
    assert ".pdf" in files[1]["errors"][0]


def test_batch_reports_per_file_and_aggregate_counts(client):
    payload = FIXTURE.read_bytes()
    response = _batch(client, [("first.xlsx", payload, None), ("second.xlsx", payload, None)], persist="true")
    body = response.json()
    assert [item["records_total"] for item in body["files"]] == [5, 5]
    assert [item["records_processed"] for item in body["files"]] == [4, 4]
    assert body["records_total"] == 10
    assert body["records_processed"] == 8
    # Counts survive persistence, so the batch view never recomputes them.
    assert client.get(f"/api/v1/batches/{body['batch_id']}").json()["records_total"] == 10


def test_run_exposes_its_batch_and_standalone_run_does_not(client):
    payload = FIXTURE.read_bytes()
    batch = _batch(client, [("first.xlsx", payload, None), ("second.xlsx", payload, None)], persist="true").json()
    child = batch["files"][0]["execution_id"]
    linked = client.get(f"/api/v1/runs/{child}/batch")
    assert linked.status_code == 200
    assert linked.json()["batch_id"] == batch["batch_id"]
    assert [item["filename"] for item in linked.json()["files"]] == ["first.xlsx", "second.xlsx"]

    standalone = _upload(client, persist="true").json()["execution_id"]
    assert client.get(f"/api/v1/runs/{standalone}/batch").status_code == 404


def test_csv_upload_produces_the_same_records_as_the_workbook(client):
    """XLSX and CSV are two encodings of one tabular contract (ADR-002)."""
    xlsx = _upload(client).json()
    csv_bytes = _fixture_as_csv()
    from_csv = _upload(client, filename="sample.csv", content=csv_bytes)
    assert from_csv.status_code == 200
    body = from_csv.json()
    assert body["records_total"] == xlsx["records_total"]
    assert body["records_processed"] == xlsx["records_processed"]
    assert body["status"] == xlsx["status"]
    assert [record["record_ref"] for record in body["records"]] == [record["record_ref"] for record in xlsx["records"]]
    assert [record["decision"] for record in body["records"]] == [record["decision"] for record in xlsx["records"]]
    assert [record["roi"]["value"] for record in body["records"]] == [record["roi"]["value"] for record in xlsx["records"]]


def test_upload_preserves_the_submitted_filename(client):
    body = _upload(client, filename="Supplier Q3 catalog.xlsx", persist="true").json()
    assert body["input_filename"] == "Supplier Q3 catalog.xlsx"


def test_upload_rejects_path_traversal_and_unsupported_types(client):
    assert _upload(client, filename="../../etc/passwd.xlsx").status_code == 200  # traversal stripped, still a workbook
    unsupported = _upload(client, filename="payload.exe")
    assert unsupported.status_code == 422
    assert ".exe" in unsupported.json()["detail"]


def test_valid_upload_runs_the_real_pipeline(client):
    resp = _upload(client)
    assert resp.status_code == 200
    body = resp.json()
    # Same fixture, same thresholds as test_pipeline_end_to_end.py -> same counts.
    assert body["records_total"] == 5
    assert body["records_processed"] == 4
    assert body["status"] == "PARTIAL_SUCCESS"


def test_response_contains_execution_id(client):
    resp = _upload(client)
    body = resp.json()
    assert body["execution_id"]
    import uuid

    uuid.UUID(body["execution_id"])  # does not raise -> a real UUID4


def test_result_matches_core_calculation(client):
    resp = _upload(client)
    records = resp.json()["records"]
    sup001 = next(r for r in records if r["record_ref"].endswith("SUP-001"))
    # Same fixture/thresholds/fees as test_pipeline_end_to_end.py::test_valid_record_gets_correct_profitability:
    # selling_price=19.99, fees=3+2=5, seller_proceeds=14.99, cost=cog5+shipping1=6 -> profit=8.99
    assert sup001["profit"]["value"] == "8.99"
    assert sup001["profit"]["status"] == "VERIFIED"


def test_result_never_flattens_field_value_to_a_bare_value(client):
    """ADR-003/ADR-004: value and verification_status must travel together."""
    resp = _upload(client)
    record = resp.json()["records"][0]
    assert {"value", "status", "provenance"}.issubset(record["asin"])
    assert record["asin"]["provenance"]["source"]
    assert record["asin"]["status"] in {"VERIFIED", "NOT_FOUND", "INFERRED", "INVALID"}


def test_single_record_endpoint_returns_canonical_snapshot_and_scopes_404s(client):
    response = _upload(client, persist="true")
    body = response.json()
    expected = next(record for record in body["records"] if record["record_ref"].endswith("SUP-001"))
    execution_id = body["execution_id"]

    found = client.get(f"/api/v1/runs/{execution_id}/records/{expected['record_ref']}")
    assert found.status_code == 200
    assert found.json() == expected

    assert client.get(f"/api/v1/runs/{execution_id}/records/not-present").status_code == 404
    assert client.get("/api/v1/runs/not-present/records/row_1").status_code == 404


def test_missing_data_record_never_invents_asin_or_price(client):
    resp = _upload(client)
    records = resp.json()["records"]
    sup002 = next(r for r in records if r["record_ref"].endswith("SUP-002"))
    assert sup002["asin"]["value"] is None
    assert sup002["asin"]["status"] == "NOT_FOUND"
    assert sup002["selling_price"]["value"] is None


def test_disqualifying_hazmat_risk_is_pass_with_reason(client):
    resp = _upload(client)
    records = resp.json()["records"]
    sup004 = next(r for r in records if r["record_ref"].endswith("SUP-004"))
    assert sup004["decision"] == "PASS"
    assert any("RISK_ABOVE_MAXIMUM" in reason for reason in sup004["decision_reasons"])
    assert sup004["hazmat_severity"] == "HIGH"  # unchanged, still provisional (ADR-010)


def test_thresholds_are_required(client):
    resp = client.post(
        "/api/v1/runs",
        files={"file": ("s.xlsx", FIXTURE.read_bytes())},
        data={"fees": _FEES},  # no thresholds field at all
    )
    assert resp.status_code == 422


def test_thresholds_missing_field_is_rejected(client):
    resp = _upload(client, thresholds=json.dumps({"target_profit": "5"}))
    assert resp.status_code == 422


def test_fees_are_required(client):
    resp = client.post(
        "/api/v1/runs",
        files={"file": ("s.xlsx", FIXTURE.read_bytes())},
        data={"thresholds": _THRESHOLDS},  # no fees field at all
    )
    assert resp.status_code == 422


def test_fees_invariant_violation_is_rejected_not_a_500(client):
    # referral_fee_rate must be in [0, 1) -- FeeInputs.__post_init__ enforces this.
    bad_fees = json.dumps({"referral_fee": "3", "referral_fee_rate": "1.5"})
    resp = _upload(client, fees=bad_fees)
    assert resp.status_code == 422


def test_invalid_excel_file_returns_422_not_500(client):
    resp = _upload(client, filename="bad.xlsx", content=b"not a real xlsx file")
    assert resp.status_code == 422
    assert "not a valid" in resp.json()["detail"]


def test_fatal_import_returns_422_with_failed_status(client, tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Supplier SKU", "Marketplace", "ASIN"])  # no "cost" column -> fatal
    ws.append(["X-1", "US", "B0TESTAAAA"])
    path = tmp_path / "broken.xlsx"
    wb.save(path)

    resp = _upload(client, content=path.read_bytes())
    assert resp.status_code == 422
    body = resp.json()
    assert body["status"] == "FAILED"
    assert body["execution_id"]


def test_no_traceback_is_ever_exposed(client):
    resp = _upload(client, filename="bad.xlsx", content=b"garbage")
    assert "Traceback" not in resp.text
    assert "site-packages" not in resp.text
    assert str(FIXTURE.parent.parent.parent) not in resp.text  # no server filesystem path leaked


def test_execution_run_not_persisted_by_default(client, tmp_path):
    resp = _upload(client)  # no persist field at all
    body = resp.json()
    assert body["persisted"] is False
    store = SqliteExecutionRunStore(tmp_path / "execution_runs.db")
    assert store.load_execution_run(body["execution_id"]) is None


def test_execution_run_persisted_when_requested(client, tmp_path):
    resp = _upload(client, persist="true")
    body = resp.json()
    assert body["persisted"] is True
    store = SqliteExecutionRunStore(tmp_path / "execution_runs.db")
    saved = store.load_execution_run(body["execution_id"])
    assert saved is not None
    assert saved.status.value == "PARTIAL_SUCCESS"
    assert saved.input_hash == body["input_hash"]


def test_download_returns_the_generated_excel(client):
    resp = _upload(client)
    execution_id = resp.json()["execution_id"]

    download = client.get(f"/api/v1/runs/{execution_id}/download")
    assert download.status_code == 200
    assert download.headers["content-type"].startswith("application/vnd.openxmlformats")

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(download.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) == 1 + 4  # header + 4 processed records, same as test_cli.py


def test_unknown_execution_id_returns_404(client):
    resp = client.get("/api/v1/runs/00000000-0000-0000-0000-000000000000/download")
    assert resp.status_code == 404


def test_uploaded_input_file_is_deleted_after_processing(client, tmp_path):
    resp = _upload(client)
    execution_id = resp.json()["execution_id"]
    run_dir = tmp_path / "juval_runs" / execution_id
    assert not (run_dir / "input.xlsx").exists()
    assert (run_dir / "output.xlsx").exists()


# -- GET /api/v1/runs/{execution_id}/records (ADR-019) ------------------


# -- GET /api/v1/runs/{execution_id} (Run Detail, ADR-019 surface) ------


def test_get_run_returns_the_persisted_summary(client):
    resp = _upload(client, persist="true")
    body = resp.json()

    detail_resp = client.get(f"/api/v1/runs/{body['execution_id']}")

    assert detail_resp.status_code == 200
    detail = detail_resp.json()
    assert detail["execution_id"] == body["execution_id"]
    assert detail["status"] == body["status"]
    assert detail["input_filename"] == body["input_filename"]
    assert detail["records_total"] == body["records_total"]
    assert detail["records_processed"] == body["records_processed"]
    assert detail["records_successful"] == body["records_successful"]
    assert detail["records_with_errors"] == body["records_with_errors"]
    assert detail["warnings"] == body["warnings"]
    assert "started_at" in detail and "finished_at" in detail


def test_get_run_unknown_execution_id_returns_404(client):
    resp = client.get("/api/v1/runs/00000000-0000-0000-0000-000000000000")
    assert resp.status_code == 404


def test_get_run_requires_a_configured_store(monkeypatch, tmp_path):
    monkeypatch.setenv("JUVAL_RUN_STORAGE_DIR", str(tmp_path))
    monkeypatch.delenv("JUVAL_EXECUTION_DB_PATH", raising=False)
    monkeypatch.delenv("JUVAL_EXECUTION_STORE", raising=False)
    unconfigured_client = TestClient(app)

    resp = unconfigured_client.get("/api/v1/runs/does-not-matter")

    assert resp.status_code == 500


def test_get_run_records_after_persist_returns_the_same_records(client):
    resp = _upload(client, persist="true")
    body = resp.json()

    records_resp = client.get(f"/api/v1/runs/{body['execution_id']}/records")

    assert records_resp.status_code == 200
    records_body = records_resp.json()
    assert records_body["execution_id"] == body["execution_id"]
    assert records_body["records"] == body["records"]


def test_get_run_records_unknown_execution_id_returns_404(client):
    resp = client.get("/api/v1/runs/00000000-0000-0000-0000-000000000000/records")
    assert resp.status_code == 404


def test_records_include_title_brand_category_dimensions(client):
    """P0 Record Intelligence (PRODUCT_CAPABILITY_MATRIX.md §3): these were
    silently dropped between the domain and the API/persisted snapshot --
    confirm they now survive both the live-upload path and the
    persist-then-reload path, with real values from the fixture."""

    resp = _upload(client, persist="true")
    body = resp.json()
    live_record = next(r for r in body["records"] if r["record_ref"].endswith("SUP-001"))

    for field_name, expected in (
        ("title", "JUVAL TEST WIDGET ALPHA"),
        ("brand", "JuvalTestBrand"),
        ("category", "Home"),
        ("height", "4"),
        ("width", "3"),
        ("length", "2"),
    ):
        assert live_record[field_name]["value"] == expected, field_name
        assert live_record[field_name]["status"] == "VERIFIED", field_name

    records_resp = client.get(f"/api/v1/runs/{body['execution_id']}/records")
    reloaded_record = next(
        r for r in records_resp.json()["records"] if r["record_ref"].endswith("SUP-001")
    )
    assert reloaded_record["title"] == live_record["title"]
    assert reloaded_record["height"] == live_record["height"]


def test_old_snapshot_missing_new_fields_still_loads(client, tmp_path):
    """Backward compatibility: a snapshot persisted before title/brand/
    category/height/width/length existed in the contract has none of those
    keys in its stored JSON. Loading it must not fail -- the new fields
    must come back as {value: None, status: None}, never a 500."""

    old_shape_snapshot = {
        "record_ref": "row_OLD-001",
        "marketplace": "US",
        "supplier_sku": "OLD-001",
        "asin": {"value": "B0OLDSHAPE1", "status": "VERIFIED"},
        "upc": {"value": None, "status": "NOT_FOUND"},
        "weight": {"value": "1.5", "status": "VERIFIED"},
        "selling_price": {"value": "19.99", "status": "VERIFIED"},
        "cog": "5",
        "shipping_per_unit": "1",
        "profit": {"value": "8.99", "status": "VERIFIED"},
        "roi": {"value": "1.5", "status": "VERIFIED"},
        "margin": {"value": "0.45", "status": "VERIFIED"},
        "break_even_price": {"value": "10", "status": "VERIFIED"},
        "max_cog_target_profit": {"value": "20", "status": "VERIFIED"},
        "max_cog_target_roi": {"value": "20", "status": "VERIFIED"},
        "hazmat_status": "ABSENT",
        "hazmat_severity": None,
        "bulky_status": "ABSENT",
        "bulky_severity": None,
        "decision": "BUY",
        "decision_reasons": [],
        "issue_count": 0,
        "issues": [],
        # no title/brand/category/height/width/length -- pre-P0 shape.
    }
    store = SqliteExecutionRunStore(tmp_path / "execution_runs.db")
    run = _run_without_records()
    store.save_execution_run(run, records=[old_shape_snapshot])

    resp = client.get(f"/api/v1/runs/{run.execution_id}/records")

    assert resp.status_code == 200
    record = resp.json()["records"][0]
    assert record["asin"]["value"] == "B0OLDSHAPE1"
    assert record["asin"]["status"] == "VERIFIED"
    assert record["asin"]["provenance"] is None
    for field_name in ("title", "brand", "category", "height", "width", "length"):
        assert record[field_name]["value"] is None, field_name
        assert record[field_name]["status"] is None, field_name


def test_get_run_records_for_run_with_no_persisted_records_returns_empty_list(client, tmp_path):
    store = SqliteExecutionRunStore(tmp_path / "execution_runs.db")
    store.save_execution_run(_run_without_records())

    resp = client.get(f"/api/v1/runs/{_run_without_records().execution_id}/records")

    assert resp.status_code == 200
    assert resp.json()["records"] == []


def test_get_run_records_requires_a_configured_store(monkeypatch, tmp_path):
    monkeypatch.setenv("JUVAL_RUN_STORAGE_DIR", str(tmp_path))
    monkeypatch.delenv("JUVAL_EXECUTION_DB_PATH", raising=False)
    monkeypatch.delenv("JUVAL_EXECUTION_STORE", raising=False)
    unconfigured_client = TestClient(app)

    resp = unconfigured_client.get("/api/v1/runs/does-not-matter/records")

    assert resp.status_code == 500


def _run_without_records():
    from datetime import datetime, timezone

    from juval.domain.execution_run import ExecutionRun, ExecutionStatus

    return ExecutionRun(
        execution_id="pre-existing-run",
        started_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
        finished_at=datetime(2026, 1, 1, 0, 5, tzinfo=timezone.utc),
        status=ExecutionStatus.SUCCESS,
        input_filename="in.xlsx",
        input_hash="abc123",
        application_version="0.1.0",
        records_total=0,
        records_processed=0,
        records_successful=0,
        records_with_errors=0,
        warnings=0,
    )


# -- GET /api/v1/runs (ADR-019) ------------------------------------------


def test_get_runs_list_is_empty_when_no_runs_persisted(client):
    resp = client.get("/api/v1/runs")
    assert resp.status_code == 200
    assert resp.json()["items"] == []


def test_get_runs_list_returns_newest_first(client):
    first = _upload(client, persist="true").json()
    second = _upload(client, persist="true").json()

    resp = client.get("/api/v1/runs")

    assert resp.status_code == 200
    execution_ids = [item["execution_id"] for item in resp.json()["items"]]
    assert execution_ids[0] == second["execution_id"]
    assert execution_ids[1] == first["execution_id"]


def test_get_runs_list_respects_limit(client):
    for _ in range(3):
        _upload(client, persist="true")

    resp = client.get("/api/v1/runs", params={"limit": 2})

    assert resp.status_code == 200
    assert len(resp.json()["items"]) == 2


def test_get_runs_list_rejects_limit_over_the_explicit_max(client):
    resp = client.get("/api/v1/runs", params={"limit": 1000})
    assert resp.status_code == 422


def test_get_runs_list_requires_a_configured_store(monkeypatch, tmp_path):
    monkeypatch.setenv("JUVAL_RUN_STORAGE_DIR", str(tmp_path))
    monkeypatch.delenv("JUVAL_EXECUTION_DB_PATH", raising=False)
    monkeypatch.delenv("JUVAL_EXECUTION_STORE", raising=False)
    unconfigured_client = TestClient(app)

    resp = unconfigured_client.get("/api/v1/runs")

    assert resp.status_code == 500
