"""Integration tests for the CLI entrypoint (interfaces/cli/main.py).

Drives `main()` directly (no subprocess) with argv lists, exactly as a
shell invocation would supply them, against the real fixture file --
consistent with test_pipeline_end_to_end.py's thresholds/fees so the
expected counts are known.
"""

from __future__ import annotations

import uuid
from pathlib import Path

import openpyxl
import pytest

from juval.infrastructure.excel.exporter import HEADERS
from juval.infrastructure.logging.sqlite_execution_run_store import SqliteExecutionRunStore
from juval.interfaces.cli.main import main

FIXTURE = Path(__file__).parent.parent / "fixtures" / "sample_sourcing_TEST_DATA.xlsx"

_BASE_ARGS = [
    "--target-profit", "5",
    "--target-roi", "0.3",
    "--max-risk-severity", "LOW",
    "--referral-fee", "3",
    "--referral-fee-rate", "0.15",
    "--fulfillment-fee", "2",
]


def test_cli_run_writes_output_and_returns_zero(tmp_path, capsys):
    output = tmp_path / "out.xlsx"
    exit_code = main([str(FIXTURE), str(output), *_BASE_ARGS, "--execution-id", "cli-test-1"])

    assert exit_code == 0
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == HEADERS
    assert len(rows) == 1 + 4  # header + 4 processed records (SUP-005 dropped, see test_pipeline_end_to_end)

    captured = capsys.readouterr()
    assert "execution_id=cli-test-1" in captured.out
    assert "status=PARTIAL_SUCCESS" in captured.out
    assert "output written to" in captured.out


def test_cli_generates_execution_id_when_not_given(tmp_path, capsys):
    output = tmp_path / "out.xlsx"
    exit_code = main([str(FIXTURE), str(output), *_BASE_ARGS])

    assert exit_code == 0
    captured = capsys.readouterr()
    printed_id = captured.out.splitlines()[0].split("execution_id=")[1].split(" ")[0]
    uuid.UUID(printed_id)  # does not raise -> looks like a real UUID4


def test_cli_persist_db_saves_execution_run_only_when_flag_given(tmp_path, capsys):
    output = tmp_path / "out.xlsx"
    db_path = tmp_path / "runs.db"

    exit_code = main(
        [str(FIXTURE), str(output), *_BASE_ARGS, "--execution-id", "cli-persist-test", "--persist-db", str(db_path)]
    )

    assert exit_code == 0
    store = SqliteExecutionRunStore(db_path)
    saved = store.load_execution_run("cli-persist-test")
    assert saved is not None
    assert saved.status.value == "PARTIAL_SUCCESS"

    captured = capsys.readouterr()
    assert f"execution_run persisted to {db_path}" in captured.out


def test_cli_without_persist_db_does_not_create_a_db_file(tmp_path):
    output = tmp_path / "out.xlsx"
    would_be_db = tmp_path / "should_not_exist.db"

    main([str(FIXTURE), str(output), *_BASE_ARGS])

    assert not would_be_db.exists()


def test_cli_missing_input_file_returns_two_without_traceback(tmp_path, capsys):
    output = tmp_path / "out.xlsx"
    missing = tmp_path / "does_not_exist.xlsx"

    exit_code = main([str(missing), str(output), *_BASE_ARGS])

    assert exit_code == 2
    assert not output.exists()
    captured = capsys.readouterr()
    assert "input file not found" in captured.err


def test_cli_missing_required_threshold_exits_nonzero(tmp_path):
    output = tmp_path / "out.xlsx"
    with pytest.raises(SystemExit) as exc_info:
        main([str(FIXTURE), str(output), "--target-roi", "0.3", "--max-risk-severity", "LOW",
              "--referral-fee", "3", "--referral-fee-rate", "0.15"])
    assert exc_info.value.code != 0


def test_cli_fatal_import_returns_one_and_writes_no_output(tmp_path, capsys):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Supplier SKU", "Marketplace", "ASIN"])  # no "cost" column -> fatal
    ws.append(["X-1", "US", "B0TESTAAAA"])
    broken = tmp_path / "broken.xlsx"
    wb.save(broken)

    output = tmp_path / "out.xlsx"
    exit_code = main([str(broken), str(output), *_BASE_ARGS])

    assert exit_code == 1
    assert not output.exists()
    captured = capsys.readouterr()
    assert "FAILED" in captured.err
