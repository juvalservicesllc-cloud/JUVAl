from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from juval.domain.provenance import VerificationStatus
from juval.domain.risk import RiskStatus, RiskType, Severity
from juval.infrastructure.excel.importer import (
    DEFAULT_RISK_SEVERITY,
    _build_risk_flag,
    import_excel,
    normalize_header,
)

NOW = datetime(2026, 1, 1, tzinfo=timezone.utc)
FIXTURE = Path(__file__).parent.parent / "fixtures" / "sample_sourcing_TEST_DATA.xlsx"


def _by_ref(records, suffix):
    for r in records:
        if r.record_ref.endswith(suffix):
            return r
    raise AssertionError(f"no record ending with {suffix!r} in {[r.record_ref for r in records]}")


@pytest.fixture(scope="module")
def import_result():
    return import_excel(FIXTURE, now=NOW)


def test_import_is_not_fatal(import_result):
    assert import_result.fatal is False


def test_five_rows_scanned_four_records_built(import_result):
    assert import_result.rows_scanned == 5
    assert len(import_result.records) == 4


def test_unknown_column_produces_warning(import_result):
    codes = [i.code for i in import_result.issues]
    assert "UNKNOWN_COLUMN" in codes


def test_valid_row_is_fully_verified(import_result):
    record = _by_ref(import_result.records, "SUP-001")
    ident = record.product.identification
    assert ident.asin.value == "B0TESTAAA1"
    assert ident.asin.status == VerificationStatus.VERIFIED
    assert ident.upc.status == VerificationStatus.VERIFIED
    assert record.product.dimensions.weight.value == Decimal("1.5")
    assert record.product.dimensions.weight.unit == "lb"
    assert record.costs is not None
    assert record.costs.cog == 5
    assert record.product.price.selling_price_used.value == Decimal("19.99")
    assert record.has_record_errors is False


def test_missing_data_row_is_not_found_not_invented(import_result):
    record = _by_ref(import_result.records, "SUP-002")
    assert record.product.identification.asin.status == VerificationStatus.NOT_FOUND
    assert record.product.identification.asin.value is None
    assert record.product.price.selling_price_used is None  # no price observed at all


def test_invalid_row_flags_are_preserved_with_raw_values(import_result):
    record = _by_ref(import_result.records, "SUP-003")
    ident = record.product.identification
    assert ident.asin.status == VerificationStatus.INVALID
    assert ident.asin.raw_value == "NOT-AN-ASIN!"
    assert ident.upc.status == VerificationStatus.INVALID
    assert record.product.dimensions.weight.status == VerificationStatus.INVALID
    assert record.costs is None  # missing COG -> no fabricated CostInputs
    assert record.has_record_errors is True
    codes = [i.code for i in record.issues]
    assert "INVALID_ASIN_FORMAT" in codes
    assert "INVALID_UPC_FORMAT" in codes
    assert "MISSING_REQUIRED_FIELD" in codes
    assert "INVALID_BOOLEAN" in codes


def test_risk_row_has_hazmat_present(import_result):
    record = _by_ref(import_result.records, "SUP-004")
    flag = record.risk.flag_for(RiskType.HAZMAT)
    assert flag.status == RiskStatus.PRESENT
    assert flag.verification_status == VerificationStatus.VERIFIED


def test_malformed_row_missing_marketplace_is_dropped(import_result):
    refs = [r.record_ref for r in import_result.records]
    assert not any("SUP-005" in ref for ref in refs)
    matching = [i for i in import_result.issues if i.record_ref and "SUP-005" in i.record_ref]
    assert any(i.code == "MISSING_REQUIRED_FIELD" and i.field == "marketplace" for i in matching)


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("ASIN", "asin"),
        ("Supplier SKU", "supplier_sku"),
        ("  Weight (lb) ", "weight_lb"),
        ("Cost-USD", "cost_usd"),
        (None, ""),
    ],
)
def test_normalize_header(raw, expected):
    assert normalize_header(raw) == expected


def test_missing_required_column_is_fatal(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Supplier SKU", "Marketplace", "ASIN"])  # no "cost" column at all
    ws.append(["X-1", "US", "B0TESTAAAA"])
    path = tmp_path / "missing_cost_column.xlsx"
    wb.save(path)

    result = import_excel(path, now=NOW)
    assert result.fatal is True
    assert result.records == ()
    assert any(i.code == "MISSING_REQUIRED_COLUMN" and i.field == "cost" for i in result.issues)


def test_empty_file_is_fatal(tmp_path):
    wb = openpyxl.Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    result = import_excel(path, now=NOW)
    assert result.fatal is True


def test_column_order_is_irrelevant(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Deliberately reversed/shuffled header order vs. the fixture.
    ws.append(["Cost", "ASIN", "Marketplace", "Supplier SKU"])
    ws.append([2.5, "B0TESTSHUF", "US", "X-2"])
    path = tmp_path / "shuffled.xlsx"
    wb.save(path)

    result = import_excel(path, now=NOW)
    assert result.fatal is False
    assert len(result.records) == 1
    record = result.records[0]
    assert record.product.identification.asin.value == "B0TESTSHUF"
    assert record.costs.cog == 2.5


def test_blank_rows_are_skipped(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Supplier SKU", "Marketplace", "ASIN", "Cost"])
    ws.append(["X-1", "US", "B0BLANKTST", 1])
    ws.append([None, None, None, None])
    ws.append([None, None, None, None])
    path = tmp_path / "with_blanks.xlsx"
    wb.save(path)

    result = import_excel(path, now=NOW)
    assert result.rows_skipped_blank == 2
    assert result.rows_scanned == 1
    assert len(result.records) == 1


# -- DEFAULT_RISK_SEVERITY fail-closed fallback (ADR-015) -------------------


def test_default_risk_severity_hazmat_still_high():
    flag = _build_risk_flag(
        {"hazmat": True}, "hazmat", RiskType.HAZMAT, source="s", now=NOW, row_ref="row_1", issues=[]
    )
    assert flag.severity == Severity.HIGH


def test_default_risk_severity_bulky_still_medium():
    flag = _build_risk_flag(
        {"bulky": True}, "bulky", RiskType.BULKY, source="s", now=NOW, row_ref="row_1", issues=[]
    )
    assert flag.severity == Severity.MEDIUM


def test_unmapped_risk_type_raises_instead_of_defaulting_to_medium():
    # Guards the test itself against silently going stale if someone adds
    # FRAGILE to DEFAULT_RISK_SEVERITY later without updating this test.
    assert RiskType.FRAGILE not in DEFAULT_RISK_SEVERITY

    with pytest.raises(KeyError):
        _build_risk_flag(
            {"fragile": True}, "fragile", RiskType.FRAGILE, source="s", now=NOW, row_ref="row_1", issues=[]
        )
